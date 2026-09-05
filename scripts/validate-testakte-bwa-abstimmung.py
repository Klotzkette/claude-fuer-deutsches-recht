#!/usr/bin/env python3
"""Gleicht die vier Arbeitsmappen mit Journal, Nebenbüchern und Jahresvortrag ab."""

from collections import defaultdict
import csv
from decimal import Decimal
from pathlib import Path
import re

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CASE = ROOT / "testakten/steuerrecht-bwa-vergleich-nuernberg"


def number(value):
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        raise ValueError(f"Keine numerische Buchung: {value!r}")
    return Decimal(str(value)).quantize(Decimal("0.01"))


def equal(left, right, location):
    if number(left) != number(right):
        raise ValueError(f"{location}: {left} != {right}")


def csv_rows(name):
    with (CASE / name).open(encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream, delimiter=";"))


def amount(value):
    return Decimal(value.replace(".", "").replace(",", "."))


def main():
    closing = {}
    transactions = 0
    for year, prefix in ((2024, "04"), (2025, "05")):
        susa = openpyxl.load_workbook(CASE / f"{prefix}_SuSa_{year}.xlsx", data_only=True)
        bwa = openpyxl.load_workbook(CASE / f"{'02' if year == 2024 else '03'}_BWA_{year}.xlsx", data_only=True)
        for workbook in (susa, bwa):
            if workbook._external_links:
                raise ValueError("Externe Arbeitsmappenverknüpfung")
            for sheet in workbook:
                if not sheet.print_area:
                    raise ValueError(f"Druckbereich fehlt: {year}/{sheet.title}")
                for row in sheet:
                    for cell in row:
                        if cell.data_type == "e":
                            raise ValueError(f"Formelfehler: {sheet.title}/{cell.coordinate}")
        annual = {row[0]: row for row in susa["JahresSuSa"].values if isinstance(row[0], str) and re.fullmatch(r"\d{4}", row[0])}
        monthly = {(row[0], row[1]): row for row in susa["Monatskonten"].values if isinstance(row[0], int)}
        journal = defaultdict(lambda: [Decimal(0), Decimal(0)])
        groups = defaultdict(Decimal)
        for row in susa["Journal"].iter_rows(min_row=6, values_only=True):
            if not row[0]:
                continue
            group, day, month, account, debit, credit, *_ = row
            if day.year != year or day.month != month or account not in annual:
                raise ValueError(f"Journalzuordnung: {group}")
            groups[group] += number(debit) - number(credit)
            journal[month, account][0] += number(debit)
            journal[month, account][1] += number(credit)
        for group, difference in groups.items():
            equal(difference, 0, group)
        transactions += len(groups)
        closing[year] = {}
        for account, row in annual.items():
            _, _, opening_debit, opening_credit, debit, credit, end_debit, end_credit = row
            balance = number(opening_debit) - number(opening_credit)
            total_debit = total_credit = Decimal(0)
            for month in range(1, 13):
                m = monthly[month, account]
                equal(m[3], balance, f"{year}/{month}/{account}: Vortrag")
                equal(m[4], journal[month, account][0], f"{year}/{month}/{account}: Soll")
                equal(m[5], journal[month, account][1], f"{year}/{month}/{account}: Haben")
                balance += number(m[4]) - number(m[5])
                equal(m[6], number(m[4]) - number(m[5]), "Monatsbewegung")
                equal(m[7], balance, "Monatsschluss")
                total_debit += number(m[4])
                total_credit += number(m[5])
            equal(debit, total_debit, f"Jahressoll {account}")
            equal(credit, total_credit, f"Jahreshaben {account}")
            equal(number(end_debit) - number(end_credit), balance, f"Jahressaldo {account}")
            closing[year][account] = balance
        for column in (2, 4, 6):
            equal(sum(number(r[column]) for r in annual.values()), sum(number(r[column + 1]) for r in annual.values()), f"SuSa {year}: Spaltenpaar {column}")
        for row in bwa["Kontenwerte"].iter_rows(min_row=6, values_only=True):
            if row[0] not in annual:
                continue
            account = row[0]
            for month in range(1, 13):
                equal(row[month + 1], monthly[month, account][6], f"BWA-Grundlage {year}/{month}/{account}")
            equal(row[16], closing[year][account], f"BWA-Kontenschluss {account}")
        equal(bwa["Jahres BWA"]["C41"].value, -sum(value for account, value in closing[year].items() if int(account) >= 4000), "Jahresergebnis aus GuV-Konten")
        for month in range(1, 13):
            result = -sum(number(row[6]) for (m, account), row in monthly.items() if m == month and int(account) >= 4000)
            equal(bwa["BWA"].cell(41, month + 1).value, result, f"Monatsergebnis {year}/{month}")
        if year == 2025:
            for row in susa["Vortrag"].iter_rows(min_row=6, values_only=True):
                if row[0] not in annual:
                    continue
                equal(row[2], closing[2024][row[0]], "Schluss 2024")
                equal(row[2] + row[3], row[4], "Abschlussbuchung")
                equal(row[4], number(annual[row[0]][2]) - number(annual[row[0]][3]), "Eröffnung 2025")
        for filename, field, account, sign in (
            ("20_OP_Debitoren_2024_2025.csv", "Offen_Brutto_EUR", "1200", 1),
            ("20_OP_Debitoren_2024_2025.csv", "EWB_EUR", "1248", -1),
            ("21_OP_Kreditoren_2024_2025.csv", "Offen_Brutto_EUR", "3300", -1),
            ("22_Inventur_2024_2025.csv", "Buchwert_Netto_EUR", "1140", 1),
        ):
            total = sum(amount(row[field]) for row in csv_rows(filename) if row["Stichtag"].startswith(str(year)))
            equal(total, sign * closing[year][account], f"{year}/{filename}/{field}")
        susa.close()
        bwa.close()
    balance = closing[2025]["1800"]
    for row in csv_rows("19_Zahlungsjournal_2026-01.csv"):
        balance += amount(row["Kontobewegung_EUR"])
        equal(balance, amount(row["Bankstand_nach_Zeile_EUR"]), "Januar-Bankstand")
        if row["Status"] == "Nicht ausgeführt":
            equal(amount(row["Kontobewegung_EUR"]), 0, "Nicht ausgeführter Auftrag")
    print(f"BWA-Abstimmung OK: vier Arbeitsmappen, {transactions} ausgeglichene Beleggruppen, Nebenbücher und Januar-Zahlungen")


if __name__ == "__main__":
    main()
