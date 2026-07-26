#!/usr/bin/env python3
"""run-eval.py - Schlanker Eval-Harness fuer Klotzkette German Legal Skills.

Liest pro Testakte (testakten/<slug>/rubric.yaml) eine Liste von Pass/Fail-
Pruefungen ein, fuehrt sie aus und schreibt einen All-Pass-Score nach
Harvey-LAB-Vorbild.

Verwendung:
    python3 scripts/run-eval.py                          # alle Testakten
    python3 scripts/run-eval.py <slug1> <slug2>         # gezielt
    python3 scripts/run-eval.py --report                # MD-Report nach EVAL_RESULTS.md

Pruefungstypen (rubric.yaml):
- file_exists          ->  Datei existiert
- text_contains        ->  Text-Substring kommt in Datei vor
- regex_match          ->  Regex matcht in Datei
- file_count           ->  Anzahl Dateien matches glob >= N
- yaml_field_equals    ->  YAML-Frontmatter-Feld hat erwarteten Wert
- json_field_equals    ->  JSON-Feld hat erwarteten Wert
- human_review         ->  Manuell zu bewerten (in der Eval als 'skipped' gewertet)
"""
from __future__ import annotations

import argparse
from email import policy
from email.parser import BytesParser
import json
import re
import sys
import zipfile
from xml.etree import ElementTree
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
TESTAKTEN = REPO / "testakten"
MARKETPLACE = REPO / ".claude-plugin" / "marketplace.json"
PLUGIN_NAMES = {
    plugin["name"]
    for plugin in json.loads(MARKETPLACE.read_text(encoding="utf-8"))["plugins"]
}
RUBRIC_EXEMPT_SLUGS = {"formatvorlagen-paradebeispiele", "megaprompts"}
SUPPORTED_CHECK_TYPES = {
    "file_exists",
    "text_contains",
    "regex_match",
    "file_count",
    "working_file_count",
    "yaml_field_equals",
    "json_field_equals",
    "human_review",
}
LEGACY_PLACEHOLDER_IDS = {
    "r03-mindestens-3-aktenstuecke",
    "r04-fachspezifischer-check-zu-ergaenzen",
}


def resolve_case_path(akte_dir: Path, raw_path: object) -> Path:
    """Löst einen Rubrikpfad auf, ohne absolute Pfade oder Ausbrüche zuzulassen."""
    if not isinstance(raw_path, str) or not raw_path.strip():
        raise ValueError("path muss eine nicht leere Zeichenfolge sein")
    relative = Path(raw_path)
    if relative.is_absolute():
        raise ValueError("absolute Pfade sind nicht zulässig")
    target = (akte_dir / relative).resolve()
    try:
        target.relative_to(akte_dir.resolve())
    except ValueError as exc:
        raise ValueError("path verlässt den Testaktenordner") from exc
    return target


def nested_value(document: object, field_path: object) -> object:
    """Liest einen mit Punkten getrennten Feldpfad aus JSON oder YAML."""
    if not isinstance(field_path, str) or not field_path.strip():
        raise ValueError("field muss eine nicht leere Zeichenfolge sein")
    current = document
    for component in field_path.split("."):
        if isinstance(current, dict) and component in current:
            current = current[component]
            continue
        if isinstance(current, list) and component.isdigit():
            index = int(component)
            if index < len(current):
                current = current[index]
                continue
        raise KeyError(field_path)
    return current


def read_searchable_text(path: Path) -> str:
    """Liest Text- und Office-Aktenstücke für inhaltliche Rubrikchecks."""
    suffix = path.suffix.lower()
    if suffix == ".docx":
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml")
        root = ElementTree.fromstring(xml)
        return "\n".join(node.text or "" for node in root.iter() if node.tag.endswith("}t"))
    if suffix == ".pdf":
        try:
            from pypdf import PdfReader  # type: ignore
        except ImportError as exc:
            raise RuntimeError("PDF-Textprüfung benötigt pypdf aus requirements.txt") from exc
        return "\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages)
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise RuntimeError("XLSX-Textprüfung benötigt openpyxl aus requirements.txt") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return "\n".join(
                "\t".join("" if value is None else str(value) for value in row)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows(values_only=True)
            )
        finally:
            workbook.close()
    if suffix == ".eml":
        with path.open("rb") as handle:
            message = BytesParser(policy=policy.default).parse(handle)
        headers = "\n".join(
            str(message.get(header, ""))
            for header in ("From", "To", "Cc", "Date", "Subject")
        )
        body_part = message.get_body(preferencelist=("plain", "html"))
        body = body_part.get_content() if body_part else ""
        return f"{headers}\n{body}"
    return path.read_text(encoding="utf-8", errors="ignore")


@dataclass
class CheckResult:
    rubric_id: str
    check_type: str
    description: str
    passed: bool | None  # None = skipped (human_review)
    detail: str = ""


@dataclass
class AktenResult:
    slug: str
    has_rubric: bool
    checks: list[CheckResult] = field(default_factory=list)

    @property
    def all_passed(self) -> bool:
        decided = [c for c in self.checks if c.passed is not None]
        return bool(decided) and all(c.passed for c in decided)

    @property
    def stats(self) -> dict[str, int]:
        passed = sum(1 for c in self.checks if c.passed is True)
        failed = sum(1 for c in self.checks if c.passed is False)
        skipped = sum(1 for c in self.checks if c.passed is None)
        return {"passed": passed, "failed": failed, "skipped": skipped}


def load_yaml(path: Path) -> dict:
    """Lädt eine Rubrik vollständig mit dem deklarierten YAML-Parser."""
    try:
        import yaml  # type: ignore
        loaded = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        if not isinstance(loaded, dict):
            error = f"YAML-Wurzel muss ein Mapping sein, gefunden: {type(loaded).__name__}"
            return {"_load_error": error}
        return loaded
    except ImportError:
        return {"_load_error": "PyYAML aus requirements.txt ist nicht installiert"}
    except Exception as exc:
        return {"_load_error": f"{type(exc).__name__}: {exc}"}


def run_check(akte_dir: Path, check: dict) -> CheckResult:
    if not isinstance(check, dict):
        return CheckResult(
            "rubric-schema",
            "schema",
            "Rubrik-Eintrag muss ein Mapping sein.",
            False,
            f"found {type(check).__name__}: {check!r}",
        )
    cid = check.get("id", "?")
    ctype = check.get("check_type", "?")
    desc = check.get("description", "")

    def ok(detail=""):
        return CheckResult(cid, ctype, desc, True, detail)

    def fail(detail):
        return CheckResult(cid, ctype, desc, False, detail)

    def skip(detail):
        return CheckResult(cid, ctype, desc, None, detail)

    if ctype == "file_exists":
        target = resolve_case_path(akte_dir, check.get("path"))
        if target.is_file():
            return ok(f"found {target.relative_to(akte_dir)}")
        return fail(f"missing {target.relative_to(akte_dir)}")

    if ctype == "text_contains":
        target = resolve_case_path(akte_dir, check.get("path"))
        if not target.is_file():
            return fail(f"file missing: {target.relative_to(akte_dir)}")
        needle = check.get("contains")
        if not isinstance(needle, str) or not needle:
            return fail("contains muss eine nicht leere Zeichenfolge sein")
        text = read_searchable_text(target)
        return ok("substring found") if needle in text else fail(f"missing substring: {needle[:60]!r}")

    if ctype == "regex_match":
        target = resolve_case_path(akte_dir, check.get("path"))
        if not target.is_file():
            return fail(f"file missing: {target.relative_to(akte_dir)}")
        pat = check.get("pattern")
        if not isinstance(pat, str) or not pat:
            return fail("pattern muss eine nicht leere Zeichenfolge sein")
        text = read_searchable_text(target)
        if re.search(pat, text, re.MULTILINE):
            return ok(f"regex matched")
        return fail(f"regex did not match: {pat!r}")

    if ctype == "file_count":
        pattern = check.get("glob", "*")
        if not isinstance(pattern, str) or not pattern or Path(pattern).is_absolute() or ".." in Path(pattern).parts:
            return fail("glob muss ein relativer, nicht leerer Pfad ohne '..' sein")
        min_count = int(check.get("min", 1))
        if min_count < 1:
            return fail("min muss mindestens 1 sein")
        files = [path for path in akte_dir.glob(pattern) if path.is_file()]
        if len(files) >= min_count:
            return ok(f"found {len(files)} files matching {pattern}")
        return fail(f"only {len(files)} files matching {pattern} (need {min_count})")

    if ctype == "working_file_count":
        from testakte_file_filter import include_in_working_dump

        min_count = int(check.get("min", 1))
        if min_count < 1:
            return fail("min muss mindestens 1 sein")
        files = [
            path
            for path in akte_dir.rglob("*")
            if include_in_working_dump(path, akte_dir, include_gesamt_pdf=False)
        ]
        if len(files) >= min_count:
            return ok(f"found {len(files)} exportable working files")
        return fail(f"only {len(files)} exportable working files (need {min_count})")

    if ctype in {"yaml_field_equals", "json_field_equals"}:
        target = resolve_case_path(akte_dir, check.get("path"))
        if not target.is_file():
            return fail(f"file missing: {target.relative_to(akte_dir)}")
        try:
            if ctype == "json_field_equals":
                document = json.loads(target.read_text(encoding="utf-8"))
            else:
                try:
                    import yaml  # type: ignore
                except ImportError as exc:
                    raise RuntimeError("YAML-Feldprüfung benötigt PyYAML") from exc
                document = yaml.safe_load(target.read_text(encoding="utf-8"))
            actual = nested_value(document, check.get("field"))
        except (json.JSONDecodeError, KeyError, RuntimeError, ValueError) as exc:
            return fail(f"field could not be read: {exc}")
        expected = check.get("equals")
        return ok(f"field equals {expected!r}") if actual == expected else fail(
            f"field value {actual!r} != {expected!r}"
        )

    if ctype == "human_review":
        return skip(check.get("note", "manual review required"))

    return fail(f"unknown check_type: {ctype}")


def rubric_schema_checks(rubric: dict) -> list[CheckResult]:
    """Prüft Metadaten und Checkstruktur, bevor Inhalte ausgewertet werden."""
    problems: list[CheckResult] = []

    def problem(rubric_id: str, detail: str) -> None:
        problems.append(CheckResult(rubric_id, "schema", "Rubrik-Schema", False, detail))

    name = rubric.get("name")
    if not isinstance(name, str) or not name.strip():
        problem("rubric-name", "name muss eine nicht leere Zeichenfolge sein")
    plugin = rubric.get("plugin")
    if plugin not in PLUGIN_NAMES:
        problem("rubric-plugin", f"unbekanntes oder fehlendes Plugin: {plugin!r}")
    checks = rubric.get("checks")
    if not isinstance(checks, list) or not checks:
        problem("rubric-checks", "checks muss eine nicht leere Liste sein")
        return problems
    seen_ids: set[str] = set()
    for index, check in enumerate(checks, start=1):
        if not isinstance(check, dict):
            problem("rubric-check", f"Check {index} ist kein Mapping")
            continue
        check_id = check.get("id")
        if not isinstance(check_id, str) or not check_id.strip():
            problem("rubric-check-id", f"Check {index} hat keine gültige id")
        elif check_id in seen_ids:
            problem("rubric-check-id", f"doppelte id: {check_id}")
        else:
            seen_ids.add(check_id)
        check_type = check.get("check_type")
        if check_type not in SUPPORTED_CHECK_TYPES:
            problem("rubric-check-type", f"Check {index}: unbekannter Typ {check_type!r}")
        if check_id in LEGACY_PLACEHOLDER_IDS:
            problem(
                "rubric-legacy-check",
                f"Check {index}: überholter Platzhalter {check_id!r}",
            )
        if (
            check_type == "file_count"
            and isinstance(check.get("glob"), str)
            and check["glob"].lower().endswith(".md")
        ):
            problem(
                "rubric-markdown-check",
                f"Check {index}: Markdown darf kein exportiertes Aktenstück sein",
            )
        description = check.get("description")
        if isinstance(description, str) and "zu ergänzen" in description.lower():
            problem(
                "rubric-placeholder-description",
                f"Check {index}: offene Platzhalterbeschreibung",
            )
    return problems


def evaluate_akte(slug: str) -> AktenResult:
    akte_dir = TESTAKTEN / slug
    result = AktenResult(slug=slug, has_rubric=False)
    rubric_path = akte_dir / "rubric.yaml"
    if not rubric_path.is_file():
        return result
    result.has_rubric = True
    rubric = load_yaml(rubric_path)
    if rubric.get("_load_error"):
        result.checks.append(
            CheckResult(
                "rubric-yaml",
                "schema",
                "Rubrik muss valides YAML sein.",
                False,
                rubric["_load_error"],
            )
        )
        return result
    result.checks.extend(rubric_schema_checks(rubric))
    checks = rubric.get("checks", [])
    if not isinstance(checks, list) or not checks:
        return result
    for check in checks:
        try:
            result.checks.append(run_check(akte_dir, check))
        except Exception as exc:
            cid = check.get("id", "rubric-schema") if isinstance(check, dict) else "rubric-schema"
            result.checks.append(
                CheckResult(
                    cid,
                    "schema",
                    "Rubrik-Check konnte nicht ausgeführt werden.",
                    False,
                    f"{type(exc).__name__}: {exc}",
                )
            )
    return result


def render_report(results: list[AktenResult]) -> str:
    lines = ["# Eval-Results", "", f"Stand: {datetime.now(timezone.utc).isoformat(timespec='seconds')}", ""]
    total = len(results)
    with_rubric = [r for r in results if r.has_rubric]
    missing = [
        r for r in results if not r.has_rubric and r.slug not in RUBRIC_EXEMPT_SLUGS
    ]
    all_pass = [r for r in with_rubric if r.all_passed]
    lines.append(f"- Testakten gesamt: **{total}**")
    lines.append(f"- mit Rubric: **{len(with_rubric)}**")
    lines.append(f"- Rubric fehlt unzulässig: **{len(missing)}**")
    lines.append(f"- All-Pass (alle Checks bestanden, ohne human_review): **{len(all_pass)}**")
    lines.append("")
    lines.append("## Detail pro Akte (nur Akten mit Rubric)")
    lines.append("")
    lines.append("| Akte | Status | passed | failed | skipped |")
    lines.append("| --- | --- | --- | --- | --- |")
    for r in with_rubric:
        s = r.stats
        status = "PASS" if r.all_passed else "FAIL"
        lines.append(f"| `{r.slug}` | {status} | {s['passed']} | {s['failed']} | {s['skipped']} |")
    lines.append("")
    failures = [(r, c) for r in with_rubric for c in r.checks if c.passed is False]
    if missing:
        lines.append("## Fehlende Rubriken")
        lines.append("")
        for result in missing:
            lines.append(f"- `{result.slug}`")
        lines.append("")
    if failures:
        lines.append("## Fehlende Checks")
        lines.append("")
        for r, c in failures:
            lines.append(f"- `{r.slug}` :: `{c.rubric_id}` ({c.check_type}): {c.detail}")
    return "\n".join(lines) + "\n"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("slugs", nargs="*", help="optional: konkrete Testakten-Slugs")
    ap.add_argument("--report", action="store_true",
                    help="Schreibt MD-Report nach EVAL_RESULTS.md")
    ap.add_argument("--json-out", help="Schreibt JSON-Snapshot fuer compare-eval-runs.py")
    ap.add_argument("--label", default="run", help="Label fuer den JSON-Snapshot (z. B. Modellname)")
    ap.add_argument("--quiet", action="store_true", help="Nur Zusammenfassung und Fehler ausgeben")
    args = ap.parse_args()

    if args.slugs:
        slugs = args.slugs
    else:
        slugs = sorted(d.name for d in TESTAKTEN.iterdir() if d.is_dir())

    results = [evaluate_akte(s) for s in slugs]

    # Konsolen-Output
    rubric_count = sum(1 for r in results if r.has_rubric)
    pass_count = sum(1 for r in results if r.has_rubric and r.all_passed)
    fail_count = sum(1 for r in results if r.has_rubric and not r.all_passed)
    missing = [
        r.slug
        for r in results
        if not r.has_rubric and r.slug not in RUBRIC_EXEMPT_SLUGS
    ]
    print(f"Testakten: {len(results)} | mit Rubric: {rubric_count} | "
          f"All-Pass: {pass_count} | Fail: {fail_count} | Rubric fehlt: {len(missing)}")
    for slug in missing:
        print(f"  [FAIL] {slug} (rubric.yaml fehlt)")
    for r in results:
        if not r.has_rubric:
            continue
        if args.quiet and r.all_passed:
            continue
        st = r.stats
        marker = "PASS" if r.all_passed else "FAIL"
        print(f"  [{marker}] {r.slug}  ({st['passed']}/{st['passed']+st['failed']} pass, "
              f"{st['skipped']} skip)")

    if args.report:
        report = render_report(results)
        (REPO / "EVAL_RESULTS.md").write_text(report, encoding="utf-8")
        print(f"Report geschrieben: EVAL_RESULTS.md")

    if args.json_out:
        snapshot = {
            "label": args.label,
            "timestamp": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "results": [
                {
                    "slug": r.slug,
                    "has_rubric": r.has_rubric,
                    "all_passed": r.all_passed,
                    "stats": r.stats,
                    "checks": [
                        {
                            "id": c.rubric_id,
                            "type": c.check_type,
                            "passed": c.passed,
                            "detail": c.detail,
                        }
                        for c in r.checks
                    ],
                }
                for r in results
            ],
        }
        Path(args.json_out).write_text(json.dumps(snapshot, indent=2), encoding="utf-8")
        print(f"JSON-Snapshot: {args.json_out}")

    return 0 if fail_count == 0 and not missing else 1


if __name__ == "__main__":
    sys.exit(main())
